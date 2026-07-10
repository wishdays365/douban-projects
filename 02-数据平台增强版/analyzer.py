import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from collections import Counter
import json
import os
from log_config import setup_logging

# 模块级 logger
logger = setup_logging()

class MovieAnalyzer:
    def __init__(self, json_path):
        self.json_path = json_path
        self.movies = self._load_json_data()
    def _load_json_data(self):
        """加载JSON数据（修复：增强字段默认值+异常捕获）"""
        try:
            if not os.path.exists(self.json_path):
                logger.warning(f"JSON文件不存在：{self.json_path}")
                return []
            with open(self.json_path, "r", encoding="utf-8") as f:
                movies = json.load(f)
        # 修复：单独捕获JSON解析错误
        except json.JSONDecodeError as e:
            logger.warning(f"JSON解析失败：{e}")
            return []
        except Exception as e:
            logger.warning(f"加载JSON失败：{e}")
            return []
        
        # 修复：增强数据清洗，确保核心字段有默认值
        for movie in movies:
            # 数值字段默认值（避免类型错误）
            movie["rating"] = float(movie.get("rating", 0.0))
            movie["year"] = int(movie.get("year", 0))
            movie["votes"] = int(movie.get("votes", 0))
            movie["duration"] = int(movie.get("duration", 0))
            # 数组字段处理（支持字符串自动分割）
            for field in ["director", "actors", "genre", "country", "language"]:
                val = movie.get(field)
                if isinstance(val, str):
                    # 按 / 分割并去除空白
                    movie[field] = [x.strip() for x in val.split("/") if x.strip()]
                elif isinstance(val, list):
                    movie[field] = val
                else:
                    movie[field] = []
        return movies

    def _convert_to_dataframe(self):
        """将JSON转为DataFrame（方便Excel生成）"""
        if not self.movies:
            return pd.DataFrame()
        # 处理数组字段：转为字符串（Excel兼容）
        movies_flat = []
        for movie in self.movies:
            flat_movie = movie.copy()
            flat_movie["director"] = "/".join(flat_movie["director"])
            flat_movie["actors"] = "/".join(flat_movie["actors"])
            flat_movie["genre"] = "/".join(flat_movie["genre"])
            flat_movie["country"] = "/".join(flat_movie["country"])
            flat_movie["language"] = "/".join(flat_movie["language"])
            movies_flat.append(flat_movie)
        return pd.DataFrame(movies_flat)

    # ========== 考察功能2：电影类型/数量统计 ==========
    def genre_analysis(self):
        """电影类型统计（适配数组型genre字段）"""
        genre_counter = Counter()
        for movie in self.movies:
            for genre in movie["genre"]:
                if genre.strip():
                    genre_counter[genre.strip()] += 1
        return dict(genre_counter.most_common(10))

    def country_analysis(self):
        """国家/地区统计（适配数组型country字段，新增日志）"""
        country_counter = Counter()
        for movie in self.movies:
            for country in movie["country"]:
                clean_country = country.strip()
                if clean_country:
                    country_counter[clean_country] += 1
        
        country_result = dict(country_counter.most_common(10))
        logger.debug(f"国家统计结果：{country_result}")
        return country_result

    def actor_analysis(self):
        """演员出镜统计（额外功能1，适配数组型actors字段）"""
        actor_counter = Counter()
        for movie in self.movies:
            for actor in movie["actors"]:
                if actor.strip():
                    actor_counter[actor.strip()] += 1
        return dict(actor_counter.most_common(10))

    def rating_year_analysis(self):
        """评分&年份统计（修复：放宽年份过滤条件）"""
        ratings = [m["rating"] for m in self.movies if m["rating"] > 0]
        # 修复：将>1900改为>0，避免无年份数据时返回空
        years = [m["year"] for m in self.movies if m["year"] > 0]
        return {
            "rating_avg": round(sum(ratings)/len(ratings), 2) if ratings else 0,
            "rating_max": max(ratings) if ratings else 0,
            "rating_min": min(ratings) if ratings else 0,
            "year_min": min(years) if years else 0,
            "year_max": max(years) if years else 0,
            "year_avg": round(sum(years)/len(years), 0) if years else 0
        }

    # ========== Excel美化生成 ==========
    def generate_beautiful_excel(self, excel_path):
        """生成美化版Excel表格（适配JSON结构）"""
        df = self._convert_to_dataframe()
        if df.empty:
            logger.info("无数据生成Excel")
            return False

        # 保存基础Excel
        df.to_excel(excel_path, index=False, engine="openpyxl")
        
        # 美化Excel样式
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        # 样式定义
        header_font = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        # 美化表头
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = cell_alignment
            cell.border = border

        # 美化内容行（9分以上评分标红）
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = cell_alignment
                cell.border = border
                # 评分列标红（匹配JSON的rating字段）
                # 动态查找rating列索引
                rating_col_index = None
                for col in ws.iter_cols(1, ws.max_column):
                    if col[0].value == "rating":
                        rating_col_index = col[0].column
                        break
                
                if rating_col_index and cell.column == rating_col_index and isinstance(cell.value, (int, float)) and cell.value >= 9.0:
                    cell.font = Font(color="FF0000", bold=True)

        # 调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(excel_path)
        logger.info(f"美化版Excel已生成：{excel_path}")
        return True

# 对外暴露的函数（适配main.py调用）
def generate_beautiful_movie_table(json_path, excel_path):
    analyzer = MovieAnalyzer(json_path)
    return analyzer.generate_beautiful_excel(excel_path)